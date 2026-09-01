"""
Monthly Attendance Summary for Arkitek MAA.

Input : one raw xPortal3000 "Daily Attendance Report (Lite)" .xlsx export
        (has Staff Name, Date, Recorded IN / OUT and an Attendance Summary per day).
Output: "ATTENDANCE REPORT_<source name>.xlsx" next to the input, containing:

    SUMMARY_MANAGEMENT   one row per employee (Chairman / Dir / Assc / Mgr)
    SUMMARY_ALLSTAFF     one row per employee (everyone else)
    DAILY                every counted day, with the computed hours - audit trail

Each summary row:
    No | Employee Name | Designation | Month/Year | Days Worked | Total Hours |
    9.00-9.15 | 9.16-9.30 | 9.31-10.00 | 10.01-10.30 | 10.31> |
    Days >= 8 Hours | Days < 8 Hours

Rules
    Counted day   working day (status not "Non-Working Day") on which the person
                  actually clocked in. Absent / no-punch days are ignored.
    Total Hours   sum of (clock-OUT minus clock-IN) over the counted days,
                  minus LUNCH_MIN per day that reaches a full shift (see below).
    Late band     by clock-IN time, only when later than 09:00.
    8-hour split  a counted day with a computable duration >= 8:00 is a
                  ">= 8 hours" day, otherwise a "< 8 hours" day.

Names, designations and the management / non-management split come from
employees.csv next to this script - keep that up to date.

Run   : double-click "Run report.bat", or  python attendance_report.py EXPORT.xlsx
Needs : Python 3.9+ and openpyxl   (pip install openpyxl)
"""
import csv
import datetime as dt
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---- rules you may want to tune ---------------------------------------------
GRACE_MIN = 9 * 60          # 09:00 - a clock-in later than this is a late arrival
FULL_DAY_MIN = 8 * 60       # a computed day of 8:00 or more is a ">= 8 hours" day
LUNCH_MIN = 60             # minutes removed from each day whose gross span is
                            # >= LUNCH_TRIGGER_MIN.  Set to 60 to deduct a
                            # 1-hour unpaid lunch;  0 = pure clock-out minus clock-in.
LUNCH_TRIGGER_MIN = 6 * 60  # only days with at least this gross span lose the lunch

BAND_LABELS = ["9.00 - 9.15", "9.16 - 9.30", "9.31 - 10.00", "10.01 - 10.30", "10.31>"]
BAND_UPPER = [15, 30, 60, 90, 10 ** 9]      # minutes after 09:00, upper edge of each band
DAYS = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}
HERE = os.path.dirname(os.path.abspath(__file__))


# ---- helpers -------------------------------------------------------------------
def norm(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def key(v):
    """Loose match key: lower-case, drop embedded ID/date digits and punctuation.

    The clock's Staff Name field sometimes carries junk ("Sahed1 - 2206",
    "Lee Wei Lek-1 010726"); stripping it lets those rows still match the master.
    Both the export names and employees.csv names go through this.
    """
    s = re.sub(r"\d+", " ", norm(v).lower())
    s = re.sub(r"[^\w@]+", " ", s)          # keep @ (e.g. "Zulkifli@Abd Latif")
    return re.sub(r"\s+", " ", s).strip()


def to_minutes(v):
    """'08:45   ' / datetime.time / '9:32' -> minutes past midnight. Blank -> 0."""
    if isinstance(v, (dt.time, dt.datetime)):
        return v.hour * 60 + v.minute
    m = re.match(r"\s*(\d{1,2}):(\d{2})", str(v or ""))
    return int(m.group(1)) * 60 + int(m.group(2)) if m else 0


def hhmm(minutes):
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def is_working_day(status):
    return not re.sub(r"[^a-z]", "", norm(status).lower()).startswith("nonwork")


def band_index(clock_in_min):
    if clock_in_min <= GRACE_MIN:
        return None
    late_by = clock_in_min - GRACE_MIN
    for i, upper in enumerate(BAND_UPPER):
        if late_by <= upper:
            return i
    return len(BAND_UPPER) - 1


def worked_minutes(clock_in, clock_out):
    """Computable day length, or None when the punch pair is unusable."""
    if clock_in <= 0 or clock_out <= clock_in:
        return None
    span = clock_out - clock_in
    if LUNCH_MIN and span >= LUNCH_TRIGGER_MIN:
        span -= LUNCH_MIN
    return span


# ---- employee master --------------------------------------------------------
def load_master(path):
    if not os.path.exists(path):
        sys.exit(f"employees.csv not found next to the script ({path}).")
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if not norm(r.get("staff_name")):
                continue
            rows.append({
                "staff_name": norm(r["staff_name"]),
                "display_name": norm(r.get("display_name")) or norm(r["staff_name"]),
                "designation": norm(r.get("designation")),
                "category": norm(r.get("category")).upper() or "STAFF",
                "active": norm(r.get("active")).upper() != "N",
                "sort": int(float(r["sort"])) if norm(r.get("sort")) else 9999,
            })
    by_key = {}
    for r in rows:
        by_key.setdefault(key(r["staff_name"]), r)
    return rows, by_key


# ---- raw xPortal export ----------------------------------------------------
# Fixed column positions in the "(Lite)" export (0-indexed). Sanity-checked below.
C_DAY, C_DATE, C_STAFFNO, C_NAME = 0, 1, 7, 9
C_IN, C_OUT, C_SUMMARY = 15, 16, 22


def _rows_from_xls(path):
    """First sheet of an old-format .xls as tuples, dates -> datetime."""
    try:
        import xlrd
    except ImportError:
        sys.exit("This export is the old .xls format. Install support with:  pip install xlrd\n"
                 "(or open it in Excel and Save As .xlsx)")
    book = xlrd.open_workbook(path)
    sh = book.sheet_by_index(0)
    for r in range(sh.nrows):
        vals = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                vals.append(xlrd.xldate.xldate_as_datetime(cell.value, book.datemode))
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                vals.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER and cell.value == int(cell.value):
                vals.append(int(cell.value))          # 18566.0 -> 18566
            else:
                vals.append(cell.value)
        yield tuple(vals)


def _raw_rows(path):
    """Yield sheet rows as tuples, for .xlsx (openpyxl) or .xls (xlrd)."""
    with open(path, "rb") as f:
        head = f.read(8)
    if head[:4] == b"PK\x03\x04":                       # zip -> real .xlsx
        ws = load_workbook(path, data_only=True).worksheets[0]
        yield from ws.iter_rows(values_only=True)
    elif head[:4] == b"\xd0\xcf\x11\xe0":               # OLE2 -> old .xls
        yield from _rows_from_xls(path)
    else:
        sys.exit("Unrecognised file - expected an Excel .xlsx or .xls export.")


def read_raw(path):
    header_seen = False
    out = []
    for row in _raw_rows(path):
        if len(row) <= C_SUMMARY:
            row = tuple(row) + (None,) * (C_SUMMARY + 1 - len(row))
        a = norm(row[C_DAY])
        if a == "Day" and norm(row[C_DATE]) == "Date":
            header_seen = True
            continue
        if a not in DAYS:
            continue
        raw_date = row[C_DATE]
        if isinstance(raw_date, dt.datetime):
            date = raw_date.date()
        else:
            m = re.match(r"\s*(\d{4})[/-](\d{1,2})[/-](\d{1,2})", str(raw_date or ""))
            if not m:
                continue
            date = dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        out.append({
            "day": a,
            "date": date,
            "staff_no": norm(row[C_STAFFNO]).rstrip("-"),
            "name": norm(row[C_NAME]),
            "in": to_minutes(row[C_IN]),
            "out": to_minutes(row[C_OUT]),
            "status": norm(row[C_SUMMARY]),
        })
    if not header_seen or not out:
        sys.exit("This does not look like an xPortal 'Daily Attendance Report (Lite)' export.")
    return out


# ---- aggregate ------------------------------------------------------------
def summarise(day_rows):
    days_worked = total_min = ge8 = lt8 = incomplete = 0
    bands = [0, 0, 0, 0, 0]
    audit = []
    for r in sorted(day_rows, key=lambda x: x["date"]):
        if not is_working_day(r["status"]):
            continue
        if r["in"] <= 0 and r["out"] <= 0:            # absent / no punch
            continue
        days_worked += 1
        bi = band_index(r["in"])
        if bi is not None:
            bands[bi] += 1
        dur = worked_minutes(r["in"], r["out"])
        if dur is None:
            incomplete += 1
            bucket = "incomplete punch"
        else:
            total_min += dur
            if dur >= FULL_DAY_MIN:
                ge8 += 1
                bucket = ">= 8h"
            else:
                lt8 += 1
                bucket = "< 8h"
        audit.append((r, dur, BAND_LABELS[bi] if bi is not None else "", bucket))
    return {"days_worked": days_worked, "total_min": total_min, "bands": bands,
            "ge8": ge8, "lt8": lt8, "incomplete": incomplete, "audit": audit}


# ---- output -------------------------------------------------------------------
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")

SUMMARY_COLS = ["No", "Employee Name", "Designation", "Month/Year", "Days Worked",
                "Total Hours"] + BAND_LABELS + ["Days >= 8 Hours", "Days < 8 Hours"]


def build_summary(wb, title, category_label, month_label, month_short, people, agg_by_key):
    ws = wb.create_sheet(title)
    ws.append(["Arkitek MAA Sdn Bhd / MAA Services Sdn Bhd"])
    ws.append(["Summary Monthly Staff Attendance Report"])
    ws.append([f"Category: {category_label}"])
    ws.append([f"Month: {month_label}"])
    ws.append([])
    ws.append(SUMMARY_COLS)
    for c in range(1, len(SUMMARY_COLS) + 1):
        ws.cell(6, c).font = BOLD
        ws.cell(6, c).alignment = CENTER
        ws.cell(6, c).fill = HEAD_FILL
    for cr in ("A1", "A2", "A3", "A4"):
        ws[cr].font = BOLD

    seq = 0
    for emp in people:
        a = agg_by_key.get(key(emp["staff_name"]))
        if a is None and not emp["active"]:
            continue
        seq += 1
        b = a["bands"] if a else [0, 0, 0, 0, 0]
        ws.append([
            seq, emp["display_name"], emp["designation"], month_short,
            a["days_worked"] if a else 0,
            hhmm(a["total_min"]) if a else "00:00",
            *(v or None for v in b),
            a["ge8"] if a else 0,
            a["lt8"] if a else 0,
        ])
    widths = [5, 34, 16, 10, 12, 12, 12, 12, 12, 12, 12, 15, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A7"
    return ws


def build_daily(wb, everyone, rows_by_key):
    ws = wb.create_sheet("DAILY")
    heads = ["Category", "Employee Name", "Designation", "Day", "Date",
             "Clock In", "Clock Out", "Worked (h:mm)", "Late Band", "8-Hour Bucket", "Status"]
    ws.append(heads)
    for c in range(1, len(heads) + 1):
        ws.cell(1, c).font = BOLD
        ws.cell(1, c).fill = HEAD_FILL
    for emp in everyone:
        a = summarise(rows_by_key.get(key(emp["staff_name"]), []))
        for r, dur, band, bucket in a["audit"]:
            ws.append([
                emp["category"], emp["display_name"], emp["designation"],
                r["day"], r["date"].strftime("%Y/%m/%d"),
                hhmm(r["in"]), hhmm(r["out"]),
                hhmm(dur) if dur is not None else "",
                band, bucket, r["status"],
            ])
    for col, w in {"A": 12, "B": 34, "C": 16, "D": 11, "E": 12, "H": 13,
                   "I": 13, "J": 14, "K": 22}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return ws


MASTER_CSV = os.path.join(HERE, "employees.csv")


# ---- orchestrate --------------------------------------------------------------
def build(src_path, lunch_min=None):
    """Build the report. Returns a results dict; does not print.

    lunch_min overrides LUNCH_MIN for this run (the UI passes it in). We set the
    module global rather than thread it through every function - fine for a
    one-report-per-process tool.
    """
    global LUNCH_MIN
    if lunch_min is not None:
        LUNCH_MIN = int(lunch_min)

    master, master_by_key = load_master(MASTER_CSV)
    records = read_raw(src_path)

    month = min(r["date"] for r in records)
    month_label = month.strftime("%B %Y")
    month_short = month.strftime("%b-%y")

    rows_by_key, unknown = {}, {}
    for rec in records:
        emp = master_by_key.get(key(rec["name"]))
        if emp is None:
            unknown.setdefault(rec["name"], rec["staff_no"])
            continue
        rows_by_key.setdefault(key(emp["staff_name"]), []).append(rec)
    agg_by_key = {k: summarise(v) for k, v in rows_by_key.items()}

    mgmt = sorted((e for e in master if e["category"] == "MGMT"), key=lambda e: e["sort"])
    staff = sorted((e for e in master if e["category"] != "MGMT"), key=lambda e: e["sort"])

    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb, "SUMMARY_MANAGEMENT",
                  "Chairman / Directors / Associates / Managers",
                  month_label, month_short, mgmt, agg_by_key)
    build_summary(wb, "SUMMARY_ALLSTAFF", "ALL",
                  month_label, month_short, staff, agg_by_key)
    build_daily(wb, mgmt + staff, rows_by_key)

    out = os.path.join(os.path.dirname(os.path.abspath(src_path)),
                       f"ATTENDANCE REPORT_{os.path.splitext(os.path.basename(src_path))[0]}.xlsx")
    wb.save(out)

    return {
        "out_path": out,
        "month_label": month_label,
        "lunch_min": LUNCH_MIN,
        "mgmt_count": sum(1 for e in mgmt if key(e["staff_name"]) in rows_by_key),
        "staff_count": sum(1 for e in staff if key(e["staff_name"]) in rows_by_key),
        "incomplete": sum(a["incomplete"] for a in agg_by_key.values()),
        "no_data": [e["display_name"] for e in master
                    if e["active"] and key(e["staff_name"]) not in rows_by_key],
        "unknown": sorted((n, sn) for n, sn in unknown.items()),
        "blank_designation": [e["display_name"] for e in master
                              if not e["designation"] and key(e["staff_name"]) in rows_by_key],
    }


def format_report(res):
    lunch = (f"{res['lunch_min']} min lunch deducted per full day" if res["lunch_min"]
             else "no lunch deduction (gross clock-out minus clock-in)")
    lines = [
        f"Month detected  : {res['month_label']}",
        f"Total Hours rule: {lunch}",
        f"Report written  : {res['out_path']}",
        f"Employees with data : {res['mgmt_count'] + res['staff_count']}  "
        f"(management {res['mgmt_count']}, staff {res['staff_count']})",
    ]
    if res["incomplete"]:
        lines.append(f"Days with a clock-in but no usable clock-out "
                     f"(left out of hours / 8h split): {res['incomplete']}")
    if res["no_data"]:
        lines += ["", "Active in employees.csv but NO rows in this export:"]
        lines += [f"  - {n}" for n in res["no_data"]]
    if res["unknown"]:
        lines += ["", "In the export but NOT in employees.csv (add them, then re-run):"]
        lines += [f"  - {n}   (staff no {sn})" for n, sn in res["unknown"]]
    if res["blank_designation"]:
        lines += ["", "No designation set in employees.csv:"]
        lines += [f"  - {n}" for n in res["blank_designation"]]
    return "\n".join(lines)


def _self_check():
    assert to_minutes("08:45   ") == 525
    assert to_minutes(dt.time(9, 5)) == 545 and to_minutes("") == 0
    assert band_index(9 * 60) is None
    assert band_index(9 * 60 + 1) == 0 and band_index(9 * 60 + 15) == 0
    assert band_index(9 * 60 + 16) == 1 and band_index(10 * 60) == 2
    assert band_index(10 * 60 + 31) == 4
    assert is_working_day("Non-Working Day") is False
    assert is_working_day("Non-WorkDay+EarlyIn+LateOut") is False
    assert is_working_day("Absent") is True
    lunch = LUNCH_MIN
    assert worked_minutes(9 * 60, 18 * 60) == 540 - lunch      # 9h gross day
    assert worked_minutes(8 * 60 + 40, 14 * 60) == 320 - (lunch if 320 >= LUNCH_TRIGGER_MIN else 0)
    assert worked_minutes(0, 18 * 60) is None
    assert worked_minutes(10 * 60, 9 * 60) is None
    d1 = worked_minutes(9 * 60 + 20, 18 * 60)                  # band2
    d2 = worked_minutes(8 * 60 + 40, 14 * 60)                  # on time, short
    s = summarise([
        {"date": dt.date(2026, 4, 1), "day": "Wednesday", "status": "LateIn+LateOut",
         "in": 9 * 60 + 20, "out": 18 * 60},
        {"date": dt.date(2026, 4, 2), "day": "Thursday", "status": "EarlyIn+EarlyOut",
         "in": 8 * 60 + 40, "out": 14 * 60},
        {"date": dt.date(2026, 4, 3), "day": "Friday", "status": "Absent",
         "in": 0, "out": 0},                                   # ignored
        {"date": dt.date(2026, 4, 4), "day": "Saturday", "status": "Non-Working Day",
         "in": 0, "out": 0},                                   # ignored
        {"date": dt.date(2026, 4, 6), "day": "Monday", "status": "LateIn",
         "in": 9 * 60 + 5, "out": 0},                          # punch, no out -> incomplete
    ])
    assert s["days_worked"] == 3, s
    assert s["bands"] == [1, 1, 0, 0, 0], s
    assert s["incomplete"] == 1, s
    assert s["total_min"] == d1 + d2, s
    assert s["ge8"] == sum(x >= FULL_DAY_MIN for x in (d1, d2)), s
    assert s["lt8"] == sum(x < FULL_DAY_MIN for x in (d1, d2)), s
    print("self-check OK")


if __name__ == "__main__":
    if len(sys.argv) == 2 and sys.argv[1] == "--self-check":
        _self_check()
        raise SystemExit

    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        try:
            import tkinter as tk
            from tkinter import filedialog
            tk.Tk().withdraw()
            path = filedialog.askopenfilename(
                title="Select the xPortal Daily Attendance (Lite) export",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        except Exception:
            pass
    if not path:
        raise SystemExit("Usage: python attendance_report.py <export.xlsx>")
    if not os.path.exists(path):
        raise SystemExit(f"File not found: {path}")
    print("\n" + format_report(build(path)) + "\n")
    try:
        if sys.stdin and sys.stdin.isatty():
            input("Press Enter to close...")
    except EOFError:
        pass
